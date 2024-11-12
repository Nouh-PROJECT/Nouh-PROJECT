import os
import re
import json
from functools import wraps
from flask import Blueprint
from flask import render_template, url_for, redirect, request, session, current_app, flash
from flask_login import login_required, current_user
from openpyxl import Workbook, load_workbook
from app.utils.db import execute_query


bp = Blueprint('quiz', __name__)

def check_quiz_owner(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if current_user.id != 1:
            query = r"SELECT u_id FROM quizzes WHERE id=%s"
            rows = execute_query(query, (kwargs.get('idx')))

            if (not rows) or (rows[0]['u_id'] != current_user.id):
                return redirect("/")
        return func(*args, **kwargs)
    return wrapper


@bp.route("/list", methods=['GET'])
@login_required
def quiz_list():
    display_type = request.args.get('type', 'authorized', type=str)
    keyword = request.args.get('keyword', type=str)

    # 페이지네이션
    page = request.args.get("page", 1, type=int)
    total_pages = 0
    start_page = 0
    end_page = 0
    per_page = 10
    offset = (page - 1) * per_page

    count_query = r"SELECT COUNT(*) num FROM quizzes"
    query = r"SELECT id, u_id, (SELECT name FROM subjects WHERE id=s_id)s, question q FROM quizzes"
    param = []

    if keyword and (display_type == "authorized"):
        query += r" WHERE u_id=%s AND question LIKE %s"
        count_query += r" WHERE u_id=%s AND question LIKE %s"
        param = [current_user.id, f"%{keyword}%"]
    elif display_type == "authorized":
        query += r" WHERE u_id=%s"
        count_query += r" WHERE u_id=%s"
        param = [current_user.id]
    elif keyword:
        query += r" WHERE question LIKE %s"
        count_query += r" WHERE question LIKE %s"
        param = [f"%{keyword}%"]

    # 페이지 계산
    rows = execute_query(count_query, tuple(param))
    total_quizzes = rows[0]['num'] if rows else 0
    total_pages = (total_quizzes + per_page - 1) // per_page

    start_page = max(page - 2, 1)
    end_page = min(page + 2, total_pages)

    # 페이지 데이터 불러오기
    query += r" LIMIT %s OFFSET %s"
    param.append(per_page)
    param.append(offset)
    quizzes = execute_query(query, tuple(param))
    quizzes = quizzes if quizzes else []

    if (end_page - start_page) < 4:
        if start_page == 1:
            end_page = min(start_page + 4, total_pages)
        elif end_page == total_pages:
            start_page = max(end_page - 4, 1)

    return render_template("quiz/list.html", quizzes=quizzes, type=display_type, page=page, total_pages=total_pages, start_page=start_page, end_page=end_page)


@bp.route("/add", methods=['GET', 'POST'])
@login_required
def quiz_add():
    query = r"SELECT id, name FROM subjects"
    results = execute_query(query)
    subjects = [(row['id'], row['name']) for row in results] if results else []

    if request.method == 'POST':
        u = current_user.id
        s = request.form.get('quiz-s')
        q = request.form.get('quiz-q')
        a = request.form.get('quiz-a')
        o1 = request.form.get('quiz-o1')
        o2 = request.form.get('quiz-o2')
        o3 = request.form.get('quiz-o3')
        c = request.form.get('quiz-c')

        if s != '0':
            query = r"INSERT INTO quizzes VALUES (NULL, %s, %s, %s, %s, %s, %s, %s, %s, NOW())"
            execute_query(query, (u,s,q,a,o1,o2,o3,c), True)
        
        return render_template("quiz/add.html", subjects=subjects)

    return render_template("quiz/add.html", subjects=subjects)


@bp.route("/detail/<idx>", methods=['GET'])
@login_required
def quiz_detail(idx: str):
    query = r"SELECT id, u_id, s_id, (SELECT name FROM subjects WHERE id=s_id)s_name, question q, "
    query += r"answer a, opt1 o1, opt2 o2, opt3 o3, comment c FROM quizzes WHERE id=%s;"
    rows = execute_query(query, (idx))
    quiz = rows[0] if rows else []

    if not quiz:
        return redirect("/")

    return render_template("quiz/detail.html", quiz=quiz)


@bp.route("/edit/<idx>", methods=['GET', 'POST'])
@login_required
@check_quiz_owner
def quiz_edit(idx: int):
    if request.method == 'POST':
        s = request.form.get('quiz-s')
        q = request.form.get('quiz-q')
        a = request.form.get('quiz-a')
        o1 = request.form.get('quiz-o1')
        o2 = request.form.get('quiz-o2')
        o3 = request.form.get('quiz-o3')
        c = request.form.get('quiz-c')

        query = r"UPDATE quizzes SET s_id=%s, question=%s, answer=%s,"
        query += r"opt1=%s, opt2=%s, opt3=%s, comment=%s WHERE id=%s"
        param = (s, q, a, o1, o2, o3, c, idx)
        execute_query(query, param, True)

    # Quiz Info
    query = r"SELECT id, u_id, s_id, question q, answer a, opt1 o1, opt2 o2, opt3 o3, comment c FROM quizzes WHERE id=%s"
    rows = execute_query(query, (idx))
    quiz = rows[0] if rows else []

    if not quiz:
        return redirect("/")

    # Subject Info
    rows = execute_query(r"SELECT id, name FROM subjects")
    subjects = [(row['id'], row['name']) for row in rows] if rows else []

    return render_template("quiz/edit.html", quiz=quiz, subjects=subjects)


@bp.route("/delete/<idx>", methods=['GET'])
@login_required
@check_quiz_owner
def quiz_delete(idx: str):
    query = r"DELETE FROM quizzes WHERE id=%s"
    execute_query(query, (idx), True)
    return redirect(url_for('quiz.quiz_list'))


@bp.route("/upload_excel", methods=['GET'])
@login_required
def quiz_upload_to_excel():
    return render_template("quiz/upload_excel.html")


@bp.route("/api/upload_excel", methods=['POST'])
@login_required
def api_upload_excel():
    file = request.files.get("file")

    if not (file and file.filename):
        return json.dumps({'status':'failed', 'message':'업로드 된 파일이 없습니다.'})

    # 파일 저장
    filepath = os.path.join(os.getcwd(), "app", "tmp", file.filename)
    file.save(filepath)

    rows = execute_query(r"SELECT id, name FROM subjects")
    subjects = {subject["name"]: subject["id"] for subject in rows} if rows else {}
    
    workbook = load_workbook(filename=filepath)
    sheets = workbook.sheetnames

    for sheet_name in sheets:
        error = False
        quizzes = []

        sheet = workbook[sheet_name]
        rows = execute_query(r"SELECT u_id FROM users WHERE name=%s", (sheet_name))
        u_id = rows[0] if rows else ""

        for i, row in enumerate(sheet.iter_rows(max_col=4, values_only=True)):
            lines = [line for line in str(row[1]).split("\n") if line.strip() != ""]
            
            if not str(row[2]).isdigit():
                error = True
                print(i, row)
                break

            if row[2] != "1":
                lines[-4], lines[int(row[2]) - 5] = lines[int(row[2]) - 5], lines[-4]

            quiz = {
                's': subjects[row[0]],
                'q': "\n".join(lines[0:-4]) if len(lines) > 5 else lines[0],
                'a': re.sub(r"^(\d|\W)[\S]*[\s]", "", lines[-4]).strip(),
                'o1': re.sub(r"^(\d|\W)[\S]*[\s]", "", lines[-3]).strip(),
                'o2': re.sub(r"^(\d|\W)[\S]*[\s]", "", lines[-2]).strip(),
                'o3': re.sub(r"^(\d|\W)[\S]*[\s]", "", lines[-1]).strip(),
                'c': row[3] if row[3] else "",
            }

            quizzes.append(quiz)

        # 에러 발생 시 예외 처리
        if error:
            workbook.close()
            return json.dumps({'status':'failed', 'message':'문제 업로드 실패!'})

        # DB에 퀴즈 추가
        # u_id = current_user.id
        rows = execute_query(r"SELECT u_id FROM users WHERE name=%s", (sheet.title))
        u_id = int(rows[0]) if rows else 1
        for quiz in quizzes:
            query = r"INSERT INTO quizzes VALUES (NULL, %s, %s, %s, %s, %s, %s, %s, %s, NULL)"
            data = (u_id, quiz['s'], quiz['q'], quiz['a'], quiz['o1'], quiz['o2'], quiz['o3'], quiz['c'])
            execute_query(query, data, True)

    workbook.close()
    return json.dumps({'status':'success', 'message':'문제 업로드 완료!'})

